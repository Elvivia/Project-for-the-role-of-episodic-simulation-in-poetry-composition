# -*- coding: utf-8 -*-


from gensim.models import KeyedVectors
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
 
def to_excel_num_2code(num):
    if isinstance(num,str): return num
    return get_column_letter(num)
def to_excel_code_2num(code):
    if isinstance(code, int): return code
    return column_index_from_string(code)

model = KeyedVectors.load('Tencent_AILab_ChineseEmbedding.bin')

filename=r'C:\Users\username\compute_file.xlsx'

filedata = openpyxl.load_workbook(filename)

filedata1 = filedata['Sheet1']

CannotfoundLine = {4:1, 17:3, 53:3, 74:1, 95:3, 107:3, 109:3, 114:1}



for i in range(2, filedata1.max_row+1):
    noun1 = filedata1.cell(row=i, column=to_excel_code_2num('A')).value.strip()
    noun2 = filedata1.cell(row=i, column=to_excel_code_2num('B')).value.strip()
    verb = filedata1.cell(row=i, column=to_excel_code_2num('C')).value.strip()
    
    try:
        simi1 = ((1-model.similarity(noun1, verb) )+(1- model.similarity(noun2, verb))) / 2
       
        filedata1.cell(row=i, column=to_excel_code_2num('E')).value = simi1
        
        
    except KeyError:
        print(f"Skipping row {i}: Word not found in the corpus.")
        continue

filedata.save(r'C:\Users\username\semantic_distance_file.xlsx')
