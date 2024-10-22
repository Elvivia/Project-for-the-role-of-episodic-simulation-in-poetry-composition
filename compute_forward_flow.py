# -*- coding: utf-8 -*-


from gensim.models import KeyedVectors
import openpyxl
from openpyxl.utils import column_index_from_string


model = KeyedVectors.load('Tencent_AILab_ChineseEmbedding.bin')

filename = r"C:\Users\username\forwardflow.xlsx"
filedata = openpyxl.load_workbook(filename)
sheet = filedata['Sheet1']

def calculate_forward_flows(verbs, noun1, noun2):
    n = len(verbs)
    if n < 2:
        return [], 0  

    instantaneous_flows = []
    total_distance = 0

    
    for i in range(1, n):
        sum_distance = 0
        for j in range(0, i):
            distance = 1 - model.similarity(verbs[i], verbs[j])
            sum_distance += distance
       
        noun_distance = ((1 - model.similarity(verbs[i], noun1)) + (1 - model.similarity(verbs[i], noun2))) / 2
        sum_distance += noun_distance
        iff = sum_distance / i  
        instantaneous_flows.append(iff)


    overall_forward_flow = sum(instantaneous_flows) / n if instantaneous_flows else 0  # 使用 n 作为除数，包含整个序列的动词总数

    return instantaneous_flows, overall_forward_flow


for i in range(2, sheet.max_row + 1):
    noun1 = sheet.cell(row=i, column=1).value
    noun2 = sheet.cell(row=i, column=2).value
    verbs = []
    col = 3  
    while True:
        cell_value = sheet.cell(row=i, column=col).value
        if cell_value is None:
            break
        verbs.append(cell_value)
        col += 1

    instantaneous_flows, overall_forward_flow = calculate_forward_flows(verbs, noun1, noun2)


    sheet.cell(row=i, column=column_index_from_string('O')).value = ','.join(map(str, instantaneous_flows))
    sheet.cell(row=i, column=column_index_from_string('P')).value = overall_forward_flow


filedata.save(r"C:\Users\username\forwardflow_updated.xlsx")
